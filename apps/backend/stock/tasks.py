import os
import logging
import csv
from datetime import datetime
import asyncio
from ftplib import FTP
from pathlib import Path
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import base64
from io import BytesIO
import zipfile

import mysql.connector
import pandas as pd
import requests
from celery import shared_task
from django.db import transaction
from django.utils import timezone
from mysql.connector import Error
from asgiref.sync import sync_to_async
from simpledbf import Dbf5

from goods.models import Product
from .models import (
    OurPriceHistory,
    OurStockSnapshot,
    Competitor,
    CompetitorBrand,
    CompetitorCategory,
    CompetitorProduct,
    CompetitorPriceStockSnapshot,
)


logger = logging.getLogger(__name__)

mysql_config = {
    "host": os.getenv("MYSQL_HOST"),
    "port": os.getenv("MYSQL_PORT"),
    "user": os.getenv("MYSQL_USER"),
    "password": os.getenv("MYSQL_PASS"),
    "database": os.getenv("MYSQL_DB"),
    "charset": os.getenv("MYSQL_CHARSET"),
}

MAX_PERCENT = Decimal("9999.99")


@shared_task
def import_our_stock_from_mysql():
    """
    Импортирует данные о складе из таблицы MySQL `our_stock` в модель OurStockSnapshot.
    """
    connection = None
    cursor = None
    total_rows = 0
    created = 0
    updated = 0
    skipped = 0

    try:
        connection = mysql.connector.connect(**mysql_config)
        if not connection.is_connected():
            logger.error("MySQL-соединение не установлено")
            return {
                "success": False,
                "error": "Не удалось установить соединение с MySQL",
            }

        cursor = connection.cursor(dictionary=True, buffered=False)

        query_parts = [
            "SELECT tovcode, reserve, fost",
            "FROM maingrey",
        ]

        query = " ".join(query_parts)
        logger.info(f"Выполняем запрос: {query}")
        cursor.execute(query)

        logger.info("Предзагружаем продукты для сопоставления ext_id -> Product...")
        products_qs = Product.objects.exclude(ext_id__isnull=True).exclude(ext_id__exact="")
        all_products = {p.ext_id: p for p in products_qs}
        logger.info(f"Загружено {len(all_products)} продуктов")

        def _to_int(value):
            if value in (None, ""):
                return 0
            try:
                if isinstance(value, Decimal):
                    return int(value)
                if isinstance(value, (int, float)):
                    return int(value)
                cleaned = str(value).replace(" ", "").replace("\xa0", "")
                if cleaned == "":
                    return 0
                return int(cleaned)
            except (ValueError, TypeError, InvalidOperation):
                return 0

        def _to_decimal_percent(value):
            if value in (None, ""):
                return None
            try:
                if isinstance(value, Decimal):
                    result = value
                else:
                    cleaned = str(value).replace(" ", "").replace("\xa0", "").replace(",", ".")
                    if cleaned == "":
                        return None
                    result = Decimal(cleaned)
                quantized = result.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                if quantized > MAX_PERCENT:
                    logger.warning(
                        "Процент %s превышает допустимый максимум %s, применяется обрезка",
                        quantized,
                        MAX_PERCENT,
                    )
                    return MAX_PERCENT
                if quantized < -MAX_PERCENT:
                    logger.warning(
                        "Процент %s меньше допустимого минимума -%s, применяется обрезка",
                        quantized,
                        MAX_PERCENT,
                    )
                    return -MAX_PERCENT
                return quantized
            except (InvalidOperation, ValueError, TypeError):
                logger.debug("Не удалось преобразовать процентное значение '%s'", value)
                return None

        def _to_decimal_value(value, quantize_pattern: str | None = "0.0001"):
            if value in (None, ""):
                return None
            try:
                if isinstance(value, Decimal):
                    result = value
                else:
                    cleaned = str(value).replace(" ", "").replace("\xa0", "").replace(",", ".")
                    if cleaned == "":
                        return None
                    result = Decimal(cleaned)
                if quantize_pattern is not None:
                    return result.quantize(Decimal(quantize_pattern), rounding=ROUND_HALF_UP)
                return result
            except (InvalidOperation, ValueError, TypeError):
                logger.debug("Не удалось преобразовать числовое значение '%s'", value)
                return None

        batch_size = 5000
        snapshot_moment = timezone.now()
        aggregated_stock: dict[str, int] = {}
        rows = cursor.fetchmany(batch_size)
        batch_num = 0
        missing_products_logged = 0
        missing_products_log_limit = 10

        while rows:
            batch_num += 1
            total_rows += len(rows)
            logger.info(
                f"Обрабатываем батч {batch_num}, записей: {len(rows)}, всего обработано: {total_rows}"
            )

            for row in rows:
                ext_id_raw = row.get("tovcode")
                if ext_id_raw is None:
                    skipped += 1
                    continue

                ext_id = str(ext_id_raw).strip()
                if not ext_id:
                    skipped += 1
                    continue

                product = all_products.get(ext_id)
                if not product:
                    if missing_products_logged < missing_products_log_limit:
                        logger.debug(
                            "Товар с ext_id=%s не найден в нашей БД, строка пропущена",
                            ext_id,
                        )
                        missing_products_logged += 1
                    skipped += 1
                    continue

                reserve_qty = _to_int(row.get("reserve"))
                fost_qty = _to_int(row.get("fost"))
                stock_qty = reserve_qty + fost_qty

                aggregated_stock[ext_id] = aggregated_stock.get(ext_id, 0) + stock_qty

            rows = cursor.fetchmany(batch_size)

        aggregated_items = list(aggregated_stock.items())
        logger.info(
            "Сформировано %s уникальных товаров для обновления остатков",
            len(aggregated_items),
        )

        logger.info("Загружаем последние процентные данные из invline...")
        cost_query = """
            SELECT il.mainbase, il.procent_up, il.procent_cust, il.ncont
            FROM invline il
            INNER JOIN (
                SELECT mainbase, MAX(id) AS last_id
                FROM invline
                WHERE mainbase IS NOT NULL 
                  AND procent_up IS NOT NULL 
                  AND procent_cust IS NOT NULL
                  AND procent_up != 0
                  AND procent_cust != 0
                GROUP BY mainbase
            ) latest ON latest.mainbase = il.mainbase 
                    AND latest.last_id = il.id
        """
        cursor.execute(cost_query)
        latest_cost_rows = cursor.fetchall()
        container_rates_map: dict[str, dict[str, Decimal | None]] = {}
        ncont_params: list[object] = []
        seen_ncont_keys: set[str] = set()

        for cost_row in latest_cost_rows:
            raw_ncont = cost_row.get("ncont")
            if raw_ncont in (None, ""):
                continue
            ncont_key = str(raw_ncont).strip()
            if not ncont_key or ncont_key in seen_ncont_keys:
                continue
            seen_ncont_keys.add(ncont_key)
            ncont_params.append(raw_ncont)

        if ncont_params:
            chunk_size = 1000
            for start in range(0, len(ncont_params), chunk_size):
                chunk = ncont_params[start : start + chunk_size]
                placeholders = ", ".join(["%s"] * len(chunk))
                container_query = (
                    f"SELECT ncont, rate, yrate FROM procont WHERE ncont IN ({placeholders})"
                )
                cursor.execute(container_query, chunk)
                container_rows = cursor.fetchall()
                for container_row in container_rows:
                    ncont_value = container_row.get("ncont")
                    if ncont_value in (None, ""):
                        continue
                    ncont_key = str(ncont_value).strip()
                    if not ncont_key:
                        continue
                    usd_rate = _to_decimal_value(container_row.get("rate"))
                    rmb_rate = _to_decimal_value(container_row.get("yrate"))
                    container_rates_map[ncont_key] = {
                        "usd_rate": usd_rate,
                        "rmb_rate": rmb_rate,
                    }

        if container_rates_map:
            logger.info("Загружено %s записей курсов контейнеров", len(container_rates_map))

        markup_cost_map: dict[str, dict[str, Decimal | None]] = {}
        for cost_row in latest_cost_rows:
            mainbase = cost_row.get("mainbase")
            if mainbase is None:
                continue
            ext_id = str(mainbase).strip()
            if not ext_id:
                continue
            markup_percent = _to_decimal_percent(cost_row.get("procent_up"))
            cost_percent = _to_decimal_percent(cost_row.get("procent_cust"))
            usd_rate = None
            rmb_rate = None
            raw_ncont = cost_row.get("ncont")
            if raw_ncont not in (None, ""):
                ncont_key = str(raw_ncont).strip()
                if ncont_key:
                    rates = container_rates_map.get(ncont_key)
                    if rates:
                        usd_rate = rates.get("usd_rate")
                        rmb_rate = rates.get("rmb_rate")
            markup_cost_map[ext_id] = {
                "markup_percent": markup_percent,
                "cost_percent": cost_percent,
                "usd_rate": usd_rate,
                "rmb_rate": rmb_rate,
            }

        logger.info(
            "Загружено %s записей процентных значений",
            len(markup_cost_map),
        )

        batch_write_size = 1000

        for start in range(0, len(aggregated_items), batch_write_size):
            chunk = aggregated_items[start : start + batch_write_size]
            with transaction.atomic():
                for ext_id, stock_qty in chunk:
                    product = all_products.get(ext_id)
                    if not product:
                        continue

                    markup_info = markup_cost_map.get(ext_id)
                    if markup_info is not None:
                        markup_percent = markup_info.get("markup_percent")
                        cost_percent = markup_info.get("cost_percent")
                        usd_rate = markup_info.get("usd_rate")
                        rmb_rate = markup_info.get("rmb_rate")
                    else:
                        markup_percent = cost_percent = usd_rate = rmb_rate = None

                    snapshot, is_created = OurStockSnapshot.objects.get_or_create(
                        product=product,
                        moment=snapshot_moment,
                        defaults={
                            "stock_qty": stock_qty,
                            "markup_percent": markup_percent,
                            "cost_percent": cost_percent,
                            "usd_rate": usd_rate,
                            "rmb_rate": rmb_rate,
                        },
                    )

                    if is_created:
                        created += 1
                        continue

                    update_fields: list[str] = []
                    if snapshot.stock_qty != stock_qty:
                        snapshot.stock_qty = stock_qty
                        update_fields.append("stock_qty")

                    if markup_info is not None and snapshot.markup_percent != markup_percent:
                        snapshot.markup_percent = markup_percent
                        update_fields.append("markup_percent")

                    if markup_info is not None and snapshot.cost_percent != cost_percent:
                        snapshot.cost_percent = cost_percent
                        update_fields.append("cost_percent")

                    if markup_info is not None and snapshot.usd_rate != usd_rate:
                        snapshot.usd_rate = usd_rate
                        update_fields.append("usd_rate")

                    if markup_info is not None and snapshot.rmb_rate != rmb_rate:
                        snapshot.rmb_rate = rmb_rate
                        update_fields.append("rmb_rate")

                    if update_fields:
                        update_fields.append("updated_at")
                        snapshot.save(update_fields=update_fields)
                        updated += 1
                    else:
                        skipped += 1

        logger.info(
            "Импорт склада завершён: всего=%s, создано=%s, обновлено=%s, пропущено=%s",
            total_rows,
            created,
            updated,
            skipped,
        )

        return {
            "success": True,
            "total": total_rows,
            "created": created,
            "updated": updated,
            "skipped": skipped,
        }

    except Error as e:
        logger.error("Ошибка при подключении к MySQL: %s", e)
        return {"success": False, "error": str(e)}
    except Exception as e:  # noqa: BLE001
        logger.error("Ошибка при импорте склада: %s", e, exc_info=True)
        return {"success": False, "error": str(e)}
    finally:
        try:
            if connection and connection.is_connected():
                if cursor:
                    cursor.close()
                connection.close()
                logger.info("MySQL соединение закрыто")
        except Exception:  # noqa: BLE001
            pass


@shared_task
def import_histprice_from_mysql(batch_size: int = 5000, from_date=None, limit=None):
    """
    Импортирует историю наших цен из таблицы MySQL `histprice` в модель OurPriceHistory.

    Ожидаемые поля в MySQL:
      - mainbase: внешний ID товара (равен Product.ext_id)
      - moment: дата/время изменения цены
      - price: цена без НДС
      - nds: ставка НДС (доля, например 0.20)
      
    Args:
        batch_size: размер батча для обработки
        from_date: загружать только записи с датой >= from_date (формат: 'YYYY-MM-DD HH:MM:SS')
        limit: максимальное количество записей для загрузки (None = все)
    """
    connection = None
    total_rows = 0
    created = 0
    updated = 0
    skipped = 0

    try:
        connection = mysql.connector.connect(**mysql_config)
        if not connection.is_connected():
            logger.error("MySQL-соединение не установлено")
            return {
                "success": False,
                "error": "Не удалось установить соединение с MySQL",
            }

        # Используем server-side cursor для больших результатов
        cursor = connection.cursor(dictionary=True, buffered=False)
        
        # Строим запрос с оптимизациями
        query_parts = [
            "SELECT mainbase, moment, price, nds",
            "FROM histprice",
            "WHERE mainbase IS NOT NULL"
        ]
        
        # Добавляем фильтр по дате если указан
        if from_date:
            query_parts.append(f"AND moment >= '{from_date}'")
            logger.info(f"Загружаем записи с даты: {from_date}")
        
        # УБРАЛИ ORDER BY чтобы не блокировать базу
        # Если нужна сортировка, добавьте индекс на поле moment в MySQL
        
        # Добавляем LIMIT если указан
        if limit:
            query_parts.append(f"LIMIT {limit}")
            logger.info(f"Ограничение записей: {limit}")
        
        query = " ".join(query_parts)
        logger.info(f"Выполняем запрос: {query}")
        
        cursor.execute(query)

        # Предзагружаем все продукты один раз для оптимизации
        logger.info("Предзагружаем все продукты для маппинга...")
        all_products = {p.ext_id: p for p in Product.objects.all()}
        logger.info(f"Загружено {len(all_products)} продуктов")

        rows = cursor.fetchmany(batch_size)
        batch_num = 0
        while rows:
            batch_num += 1
            total_rows += len(rows)
            logger.info(f"Обрабатываем батч {batch_num}, записей: {len(rows)}, всего обработано: {total_rows}")
            
            with transaction.atomic():
                # Используем предзагруженный словарь продуктов
                for r in rows:
                    ext_id = str(r.get("mainbase")) if r.get("mainbase") is not None else None
                    if not ext_id:
                        skipped += 1
                        continue

                    product = all_products.get(ext_id)
                    if not product:
                        # Нет соответствующего товара в нашей БД
                        skipped += 1
                        continue

                    # Приведение moment к datetime
                    moment_val = r.get("moment")
                    if isinstance(moment_val, str):
                        try:
                            # Попытка ISO8601
                            moment = datetime.fromisoformat(moment_val)
                        except Exception:
                            # На всякий случай универсальный парсинг
                            from dateutil import parser  # type: ignore

                            moment = parser.parse(moment_val)
                    else:
                        moment = moment_val

                    if moment is None:
                        skipped += 1
                        continue

                    # Делаем datetime timezone-aware если он naive
                    if moment.tzinfo is None or moment.utcoffset() is None:
                        moment = timezone.make_aware(moment)

                    price = r.get("price")
                    vat = r.get("nds")
                    
                    # Применяем скидку 15% к цене
                    if price is not None:
                        price_with_discount = float(price) * 0.85
                    else:
                        price_with_discount = 0

                    obj, is_created = OurPriceHistory.objects.get_or_create(
                        product=product,
                        moment=moment,
                        defaults={
                            "price_ex_vat": price_with_discount,
                            "vat_rate": vat if vat is not None else 0,
                        },
                    )
                    if is_created:
                        created += 1
                    else:
                        # Обновляем при изменении
                        changed = False
                        if price is not None and obj.price_ex_vat != price_with_discount:
                            obj.price_ex_vat = price_with_discount
                            changed = True
                        if vat is not None and obj.vat_rate != vat:
                            obj.vat_rate = vat
                            changed = True
                        if changed:
                            obj.save(update_fields=["price_ex_vat", "vat_rate", "updated_at"])
                            updated += 1
                        else:
                            skipped += 1

            rows = cursor.fetchmany(batch_size)
            
            # Логируем прогресс каждые 10 батчей
            if batch_num % 10 == 0:
                logger.info(
                    f"Прогресс: батч {batch_num}, всего={total_rows}, "
                    f"создано={created}, обновлено={updated}, пропущено={skipped}"
                )

        logger.info(
            "Импорт histprice завершён: всего=%s, создано=%s, обновлено=%s, пропущено=%s",
            total_rows,
            created,
            updated,
            skipped,
        )
        return {
            "success": True,
            "total": total_rows,
            "created": created,
            "updated": updated,
            "skipped": skipped,
        }

    except Error as e:
        logger.error(f"Ошибка при подключении к MySQL: {e}")
        return {"success": False, "error": str(e)}
    except Exception as e:  # noqa: BLE001
        logger.error(f"Ошибка при импорте histprice: {e}", exc_info=True)
        return {"success": False, "error": str(e)}
    finally:
        try:
            if connection and connection.is_connected():
                cursor.close()
                connection.close()
                logger.info("MySQL соединение закрыто")
        except Exception:  # noqa: BLE001
            pass


@shared_task
def import_prom_from_ftp():
    """
    Импортирует данные о товарах конкурента PROM из FTP.
    
    Оптимизированная версия с bulk-операциями:
    1. Подключается к FTP серверу конкурента PROM
    2. Скачивает файл Item.csv
    3. Парсит CSV и создает/обновляет записи CompetitorProduct батчами
    4. Создает снимки цен CompetitorPriceStockSnapshot батчами
    
    Возвращает статистику по импортированным данным.
    """
    from django.conf import settings
    from django.utils import timezone
    
    logger.info("Начинаем импорт PROM из FTP")
    
    # Находим конкурента PROM в базе
    try:
        competitor = Competitor.objects.get(name="PROM", data_source_type=Competitor.DataSourceType.FTP_CSV)
    except Competitor.DoesNotExist:
        error_msg = "Конкурент PROM с типом FTP не найден в базе данных"
        logger.error(error_msg)
        return {"success": False, "error": error_msg}
    
    logger.info(f"Найден конкурент: {competitor.name}")
    logger.info(f"FTP настройки: host={competitor.data_url}, user={competitor.username}")
    
    # Очищаем FTP хост от протокола
    ftp_host = competitor.data_url.strip()
    if ftp_host.startswith('ftp://'):
        ftp_host = ftp_host[6:]
    elif ftp_host.startswith('ftps://'):
        ftp_host = ftp_host[7:]
    ftp_host = ftp_host.rstrip('/')
    
    logger.info(f"Очищенный FTP хост: '{ftp_host}'")
    
    # Создаем директорию для скачивания
    base_download_dir = Path(settings.MEDIA_ROOT) / "ftp_downloads"
    competitor_dir = base_download_dir / competitor.name
    competitor_dir.mkdir(parents=True, exist_ok=True)
    
    csv_file_path = competitor_dir / "Item.csv"
    
    try:
        # Подключаемся к FTP и скачиваем файл
        logger.info("Подключаемся к FTP серверу...")
        ftp = FTP(ftp_host)
        ftp.login(user=competitor.username or 'anonymous', passwd=competitor.password or '')
        
        logger.info("Скачиваем файл Item.csv...")
        with open(csv_file_path, 'wb') as local_file:
            ftp.retrbinary('RETR Item.csv', local_file.write)
        
        ftp.quit()
        logger.info(f"Файл успешно скачан: {csv_file_path}")
        
    except Exception as e:
        error_msg = f"Ошибка при скачивании файла с FTP: {str(e)}"
        logger.error(error_msg)
        return {"success": False, "error": error_msg}
    
    # Парсим CSV файл
    try:
        logger.info("Начинаем парсинг CSV файла...")
        
        products_created = 0
        products_updated = 0
        brands_created = 0
        snapshots_created = 0
        errors = []
        
        collected_at = timezone.now()
        
        # Пробуем разные кодировки для чтения файла
        encodings_to_try = ['windows-1251', 'cp1251', 'utf-8-sig', 'utf-8', 'latin-1']
        csv_data = None
        used_encoding = None
        
        for encoding in encodings_to_try:
            try:
                with open(csv_file_path, 'r', encoding=encoding) as f:
                    # Используем точку с запятой как разделитель
                    reader = csv.DictReader(f, delimiter=';')
                    csv_data = list(reader)
                    used_encoding = encoding
                    break
            except UnicodeDecodeError:
                continue
        
        if csv_data is None:
            raise ValueError("Не удалось определить кодировку файла")
        
        logger.info(f"Файл прочитан с кодировкой: {used_encoding}")
        
        rows = csv_data
        total_rows = len(rows)
        logger.info(f"Прочитано {total_rows} строк из Item.csv")
        
        # Проверяем наличие необходимых колонок
        if rows:
            required_columns = ['ITEM_ID', 'NAME']
            available_columns = list(rows[0].keys())
            missing_columns = [col for col in required_columns if col not in available_columns]
            if missing_columns:
                error_msg = f"В CSV файле отсутствуют обязательные колонки: {missing_columns}. Доступные колонки: {available_columns}"
                logger.error(error_msg)
                return {"success": False, "error": error_msg}
            logger.info(f"Колонки CSV: {available_columns[:15]}{'...' if len(available_columns) > 15 else ''}")
        
        # Счетчики для пропущенных записей
        skipped_no_item_id = 0
        skipped_no_part_number = 0
        skipped_parse_errors = 0
        processed_rows = 0
        
        # Предзагрузка всех существующих брендов для кэширования
        logger.info("Загружаем существующие бренды...")
        existing_brands = {
            brand.name: brand 
            for brand in CompetitorBrand.objects.filter(competitor=competitor).select_related('competitor')
        }
        logger.info(f"Загружено {len(existing_brands)} существующих брендов")
        
        # Предзагрузка всех существующих продуктов для обновления
        # ИСПРАВЛЕНО: загружаем по ext_id, т.к. constraint в БД на (competitor_id, ext_id)
        logger.info("Загружаем существующие продукты...")
        existing_products_by_ext_id = {
            prod.ext_id: prod
            for prod in CompetitorProduct.objects.filter(competitor=competitor).select_related('brand', 'competitor')
        }
        existing_products_by_part = {
            prod.part_number: prod
            for prod in CompetitorProduct.objects.filter(competitor=competitor).select_related('brand', 'competitor')
        }
        logger.info(f"Загружено {len(existing_products_by_ext_id)} существующих продуктов")
        
        # Обрабатываем батчами для оптимизации
        batch_size = 2000
        batch_num = 0
        for i in range(0, total_rows, batch_size):
            batch_num += 1
            batch = rows[i:i + batch_size]
            logger.info(f"Обрабатываем батч {batch_num} ({len(batch)} строк, строки {i+1}-{min(i+len(batch), total_rows)})")
            
            # Для первого батча показываем пример структуры данных
            if batch_num == 1 and batch:
                logger.info(f"Пример первой строки CSV: {list(batch[0].keys())[:10]}...")
            
            # Коллекции для bulk операций
            brands_to_create = []
            products_to_create = []
            products_to_update = []
            snapshots_to_create = []
            
            # Временный кэш новых брендов в этом батче
            new_brands_cache = {}
            
            # Кэш для отслеживания дубликатов ext_id в текущем батче
            seen_ext_ids = set()
            
            with transaction.atomic():
                # Первый проход: собираем данные и создаем бренды
                parsed_rows = []
                for row in batch:
                    try:
                        # Извлекаем данные из CSV
                        item_id = row.get('ITEM_ID', '').strip()
                        if not item_id:
                            skipped_no_item_id += 1
                            continue
                        
                        # Проверяем дубликат ext_id в текущем батче
                        if item_id in seen_ext_ids:
                            # Пропускаем дубликаты в пределах одного батча
                            if len([err for err in errors if 'дубликат ITEM_ID' in err]) < 5:
                                errors.append(f"Пропущен дубликат ITEM_ID={item_id} в батче {batch_num}")
                            skipped_parse_errors += 1
                            continue
                        seen_ext_ids.add(item_id)
                        
                        part_number = row.get('NAME', '').strip()
                        if not part_number:
                            skipped_no_part_number += 1
                            continue
                        
                        producer = row.get('PRODUCER', '').strip()
                        
                        # Цена (формат: "1 642,42" - пробел как разделитель тысяч, запятая как десятичная)
                        price_ex_vat = None
                        price_str = ''
                        for price_field in ['PB_5', 'PB_4', 'PB_3', 'PB_2', 'PB_1']:
                            value = row.get(price_field, '').strip()
                            if value:
                                price_str = value
                                break
                        if price_str:
                            try:
                                # Убираем все пробелы и неразрывные пробелы, заменяем запятую на точку
                                price_clean = price_str.replace(' ', '').replace('\xa0', '').replace(',', '.')
                                price_ex_vat = Decimal(price_clean)
                            except (InvalidOperation, ValueError) as e:
                                # Логируем только первые 5 ошибок парсинга цены в батче для отладки
                                if len([err for err in errors if 'цены' in err]) < 5:
                                    errors.append(f"Ошибка парсинга цены '{price_str}' для ITEM_ID={item_id}: {e}")
                                pass
                        
                        # ИСПРАВЛЕНО: Количество на складе - складываем числа, а не строки
                        stock_qty = 0
                        for_sale = row.get('FOR_SALE', '0').strip()
                        for_sale2 = row.get('FOR_SALE2', '0').strip()
                        
                        try:
                            # Убираем пробелы из числа перед конвертацией
                            for_sale_clean = for_sale.replace(' ', '').replace('\xa0', '') if for_sale else '0'
                            stock_qty += int(for_sale_clean) if for_sale_clean else 0
                        except ValueError:
                            pass
                        
                        try:
                            # Убираем пробелы из числа перед конвертацией
                            for_sale2_clean = for_sale2.replace(' ', '').replace('\xa0', '') if for_sale2 else '0'
                            stock_qty += int(for_sale2_clean) if for_sale2_clean else 0
                        except ValueError:
                            pass
                        
                        # Определяем статус наличия
                        if stock_qty > 10:
                            stock_status = CompetitorPriceStockSnapshot.StockStatus.IN_STOCK
                        elif stock_qty > 0:
                            stock_status = CompetitorPriceStockSnapshot.StockStatus.LOW_STOCK
                        else:
                            stock_status = CompetitorPriceStockSnapshot.StockStatus.OUT_OF_STOCK
                        
                        # Создаем/находим бренд
                        brand = None
                        if producer:
                            # Проверяем в кэше существующих
                            if producer in existing_brands:
                                brand = existing_brands[producer]
                            # Проверяем в кэше новых в этом батче
                            elif producer in new_brands_cache:
                                brand = new_brands_cache[producer]
                            # Создаем новый бренд
                            else:
                                brand = CompetitorBrand(
                                    competitor=competitor,
                                    name=producer,
                                    ext_id=''
                                )
                                brands_to_create.append(brand)
                                new_brands_cache[producer] = brand
                        
                        # Сохраняем распарсенные данные
                        parsed_rows.append({
                            'item_id': item_id,
                            'part_number': part_number,
                            'brand': brand,
                            'producer': producer,
                            'price_ex_vat': price_ex_vat,
                            'stock_qty': stock_qty,
                            'stock_status': stock_status,
                            'row': row
                        })
                        
                    except Exception as e:
                        error_msg = f"Ошибка при парсинге строки {row.get('ITEM_ID', 'unknown')}: {str(e)}"
                        logger.error(error_msg)
                        errors.append(error_msg)
                        skipped_parse_errors += 1
                        continue
                
                processed_rows += len(parsed_rows)
                
                logger.info(f"Распарсено {len(parsed_rows)} валидных строк из {len(batch)} в батче {batch_num}")
                
                # Создаем новые бренды батчем
                # ИСПРАВЛЕНО: используем ignore_conflicts=True для обработки дубликатов
                if brands_to_create:
                    created_brand_names = [b.name for b in brands_to_create]
                    CompetitorBrand.objects.bulk_create(brands_to_create, ignore_conflicts=True)
                    brands_created += len(brands_to_create)
                    logger.info(f"Создано {len(brands_to_create)} новых брендов")
                    
                    # ВАЖНО: Перезагружаем созданные бренды из БД для получения id
                    # bulk_create с ignore_conflicts=True не возвращает id
                    freshly_created_brands = CompetitorBrand.objects.filter(
                        competitor=competitor,
                        name__in=created_brand_names
                    ).select_related('competitor')
                    
                    # Обновляем кэш существующих брендов со свежими объектами из БД
                    for brand in freshly_created_brands:
                        existing_brands[brand.name] = brand
                        # Также обновляем в new_brands_cache для текущего батча
                        if brand.name in new_brands_cache:
                            new_brands_cache[brand.name] = brand
                
                # Второй проход: создаем/обновляем продукты
                for parsed in parsed_rows:
                    try:
                        # ИСПРАВЛЕНО: проверяем по ext_id (ITEM_ID), т.к. constraint на (competitor_id, ext_id)
                        existing_product = existing_products_by_ext_id.get(parsed['item_id'])
                        
                        # Получаем бренд из кэша (с актуальным id из БД)
                        brand = None
                        if parsed['producer']:
                            brand = existing_brands.get(parsed['producer'])
                        
                        if existing_product:
                            # Обновляем существующий продукт
                            existing_product.ext_id = parsed['item_id']
                            existing_product.name = parsed['part_number']
                            existing_product.brand = brand
                            existing_product.tech_params = {
                                'body': parsed['row'].get('BODY', ''),
                                'year': parsed['row'].get('YEAR_', ''),
                                'country': parsed['row'].get('COUNTRY', ''),
                                'packname': parsed['row'].get('PACKNAME', ''),
                                'pack_quant': parsed['row'].get('PACK_QUANT', ''),
                                'weight': parsed['row'].get('WEIGHT', ''),
                                'datasheet': parsed['row'].get('DATASHEET', ''),
                                'photo_url': parsed['row'].get('PHOTO_URL', ''),
                            }
                            products_to_update.append(existing_product)
                            parsed['product'] = existing_product
                        else:
                            # Создаем новый продукт
                            new_product = CompetitorProduct(
                                competitor=competitor,
                                part_number=parsed['part_number'],
                                ext_id=parsed['item_id'],
                                name=parsed['part_number'],
                                brand=brand,
                                tech_params={
                                    'body': parsed['row'].get('BODY', ''),
                                    'year': parsed['row'].get('YEAR_', ''),
                                    'country': parsed['row'].get('COUNTRY', ''),
                                    'packname': parsed['row'].get('PACKNAME', ''),
                                    'pack_quant': parsed['row'].get('PACK_QUANT', ''),
                                    'weight': parsed['row'].get('WEIGHT', ''),
                                    'datasheet': parsed['row'].get('DATASHEET', ''),
                                    'photo_url': parsed['row'].get('PHOTO_URL', ''),
                                }
                            )
                            products_to_create.append(new_product)
                            parsed['product'] = new_product
                            
                    except Exception as e:
                        error_msg = f"Ошибка при подготовке продукта {parsed['part_number']}: {str(e)}"
                        logger.error(error_msg)
                        errors.append(error_msg)
                        continue
                
                # Bulk create новых продуктов
                # ИСПРАВЛЕНО: используем ignore_conflicts=True для обработки дубликатов
                if products_to_create:
                    created_ext_ids = [p.ext_id for p in products_to_create]
                    CompetitorProduct.objects.bulk_create(products_to_create, ignore_conflicts=True)
                    products_created += len(products_to_create)
                    logger.info(f"Создано {len(products_to_create)} новых продуктов")
                    
                    # ВАЖНО: Перезагружаем созданные продукты из БД для получения id
                    # bulk_create с ignore_conflicts=True не возвращает id
                    freshly_created = CompetitorProduct.objects.filter(
                        competitor=competitor,
                        ext_id__in=created_ext_ids
                    ).select_related('brand', 'competitor')
                    
                    # Обновляем кэш существующих продуктов со свежими объектами из БД
                    for product in freshly_created:
                        existing_products_by_ext_id[product.ext_id] = product
                        existing_products_by_part[product.part_number] = product
                
                # Bulk update существующих продуктов
                if products_to_update:
                    CompetitorProduct.objects.bulk_update(
                        products_to_update,
                        ['ext_id', 'name', 'brand', 'tech_params', 'updated_at'],
                        batch_size=1000
                    )
                    products_updated += len(products_to_update)
                    logger.info(f"Обновлено {len(products_to_update)} продуктов")
                
                # Третий проход: создаем снимки
                for parsed in parsed_rows:
                    try:
                        if 'product' not in parsed:
                            continue
                        
                        # ВАЖНО: Используем продукт из кэша (с актуальным id из БД)
                        product = existing_products_by_ext_id.get(parsed['item_id'])
                        if not product:
                            # Если по какой-то причине не нашли, пропускаем
                            logger.warning(f"Не найден продукт с ext_id={parsed['item_id']} для создания снимка")
                            continue
                        
                        snapshot = CompetitorPriceStockSnapshot(
                            competitor=competitor,
                            competitor_product=product,
                            collected_at=collected_at,
                            price_ex_vat=parsed['price_ex_vat'],
                            stock_qty=parsed['stock_qty'],
                            stock_status=parsed['stock_status'],
                            currency='RUB',
                            raw_payload=dict(parsed['row'])
                        )
                        snapshots_to_create.append(snapshot)
                        
                    except Exception as e:
                        error_msg = f"Ошибка при создании снимка для {parsed.get('part_number', 'unknown')}: {str(e)}"
                        logger.error(error_msg)
                        errors.append(error_msg)
                        continue
                
                # Bulk create снимков
                if snapshots_to_create:
                    CompetitorPriceStockSnapshot.objects.bulk_create(
                        snapshots_to_create, 
                        ignore_conflicts=True,
                        batch_size=1000
                    )
                    snapshots_created += len(snapshots_to_create)
                    logger.info(f"Создано {len(snapshots_to_create)} снимков")
            
            logger.info(
                f"Батч завершен. Всего создано: продуктов={products_created}, "
                f"обновлено={products_updated}, брендов={brands_created}, снимков={snapshots_created}"
            )
        
        result = {
            "success": True,
            "total_rows": total_rows,
            "processed_rows": processed_rows,
            "products_created": products_created,
            "products_updated": products_updated,
            "brands_created": brands_created,
            "snapshots_created": snapshots_created,
            "skipped_no_item_id": skipped_no_item_id,
            "skipped_no_part_number": skipped_no_part_number,
            "skipped_parse_errors": skipped_parse_errors,
            "errors_count": len(errors),
            "errors": errors[:10] if errors else []  # Первые 10 ошибок
        }
        
        logger.info(
            f"✅ Импорт PROM завершен успешно! "
            f"Всего строк в CSV: {total_rows}, обработано валидных: {processed_rows} ({processed_rows/total_rows*100:.1f}%). "
            f"Создано: {products_created} продуктов, {brands_created} брендов, {snapshots_created} снимков. "
            f"Обновлено: {products_updated} продуктов. "
            f"Пропущено: {skipped_no_item_id} без ITEM_ID, {skipped_no_part_number} без NAME, {skipped_parse_errors} с ошибками парсинга. "
            f"Всего ошибок: {len(errors)}"
        )
        return result
        
    except Exception as e:
        error_msg = f"Ошибка при парсинге CSV файла: {str(e)}"
        logger.error(error_msg, exc_info=True)
        return {"success": False, "error": error_msg}


@shared_task
def import_rct_from_https():
    """
    Импортирует данные о товарах конкурента RCT из HTTPS CSV файла.
    
    Оптимизированная версия с bulk-операциями:
    1. Находит конкурента RCT в базе
    2. Скачивает CSV файл по HTTPS ссылке из data_url
    3. Парсит CSV со столбцами: Номенклатура;Описание;Код;Тип корпуса;Производитель;Аналоги;Цена 4;Свободный остаток;Ожидается;Кратность отгрузки
    4. Создает/обновляет записи CompetitorProduct батчами
    5. Создает снимки цен CompetitorPriceStockSnapshot батчами
    
    Маппинг полей:
    - CompetitorProduct.part_number = Номенклатура
    - CompetitorProduct.ext_id = Код
    - CompetitorProduct.brand = Производитель
    - CompetitorProduct.name = Описание
    - CompetitorPriceStockSnapshot.price_ex_vat = Цена 4 (в USD)
    - CompetitorPriceStockSnapshot.stock_qty = Свободный остаток
    - CompetitorPriceStockSnapshot.currency = USD
    
    Возвращает статистику по импортированным данным.
    """
    from django.conf import settings
    from django.utils import timezone
    
    logger.info("Начинаем импорт RCT из HTTPS")
    
    # Находим конкурента RCT в базе
    try:
        competitor = Competitor.objects.get(name="RCT", data_source_type=Competitor.DataSourceType.HTTPS)
    except Competitor.DoesNotExist:
        error_msg = "Конкурент RCT с типом HTTPS не найден в базе данных"
        logger.error(error_msg)
        return {"success": False, "error": error_msg}
    
    if not competitor.data_url:
        error_msg = "У конкурента RCT не указан URL для скачивания данных"
        logger.error(error_msg)
        return {"success": False, "error": error_msg}
    
    logger.info(f"Найден конкурент: {competitor.name}")
    logger.info(f"URL для скачивания: {competitor.data_url}")
    
    # Создаем директорию для скачивания
    base_download_dir = Path(settings.MEDIA_ROOT) / "https_downloads"
    competitor_dir = base_download_dir / competitor.name
    competitor_dir.mkdir(parents=True, exist_ok=True)
    
    csv_file_path = competitor_dir / "data.csv"
    
    try:
        # Скачиваем файл через HTTPS
        logger.info("Скачиваем CSV файл...")
        
        # Настраиваем заголовки и аутентификацию если нужно
        headers = {}
        auth = None
        if competitor.username and competitor.password:
            auth = (competitor.username, competitor.password)
        
        response = requests.get(
            competitor.data_url,
            headers=headers,
            auth=auth,
            timeout=300,  # 5 минут таймаут
            stream=True
        )
        response.raise_for_status()
        
        # Сохраняем файл
        with open(csv_file_path, 'wb') as local_file:
            for chunk in response.iter_content(chunk_size=8192):
                local_file.write(chunk)
        
        logger.info(f"Файл успешно скачан: {csv_file_path}")
        
    except requests.exceptions.RequestException as e:
        error_msg = f"Ошибка при скачивании файла по HTTPS: {str(e)}"
        logger.error(error_msg)
        return {"success": False, "error": error_msg}
    except Exception as e:
        error_msg = f"Ошибка при сохранении файла: {str(e)}"
        logger.error(error_msg)
        return {"success": False, "error": error_msg}
    
    # Парсим CSV файл
    try:
        logger.info("Начинаем парсинг CSV файла...")
        
        products_created = 0
        products_updated = 0
        brands_created = 0
        snapshots_created = 0
        errors = []
        
        collected_at = timezone.now()
        
        # Пробуем разные кодировки для чтения файла
        encodings_to_try = ['windows-1251', 'cp1251', 'utf-8-sig', 'utf-8', 'latin-1']
        csv_data = None
        used_encoding = None
        
        for encoding in encodings_to_try:
            try:
                with open(csv_file_path, 'r', encoding=encoding) as f:
                    # Используем точку с запятой как разделитель
                    reader = csv.DictReader(f, delimiter=';')
                    csv_data = list(reader)
                    used_encoding = encoding
                    break
            except UnicodeDecodeError:
                continue
        
        if csv_data is None:
            raise ValueError("Не удалось определить кодировку файла")
        
        logger.info(f"Файл прочитан с кодировкой: {used_encoding}")
        
        rows = csv_data
        total_rows = len(rows)
        logger.info(f"Прочитано {total_rows} строк из CSV")
        
        # Проверяем наличие необходимых колонок
        if rows:
            required_columns = ['Номенклатура', 'Код']
            available_columns = list(rows[0].keys())
            missing_columns = [col for col in required_columns if col not in available_columns]
            if missing_columns:
                error_msg = f"В CSV файле отсутствуют обязательные колонки: {missing_columns}. Доступные колонки: {available_columns}"
                logger.error(error_msg)
                return {"success": False, "error": error_msg}
            logger.info(f"Колонки CSV: {available_columns}")
        
        # Счетчики для пропущенных записей
        skipped_no_code = 0
        skipped_no_part_number = 0
        skipped_parse_errors = 0
        processed_rows = 0
        
        # Предзагрузка всех существующих брендов для кэширования
        logger.info("Загружаем существующие бренды...")
        existing_brands = {
            brand.name: brand 
            for brand in CompetitorBrand.objects.filter(competitor=competitor).select_related('competitor')
        }
        logger.info(f"Загружено {len(existing_brands)} существующих брендов")
        
        # Предзагрузка всех существующих продуктов для обновления
        logger.info("Загружаем существующие продукты...")
        existing_products_by_ext_id = {
            prod.ext_id: prod
            for prod in CompetitorProduct.objects.filter(competitor=competitor).select_related('brand', 'competitor')
        }
        logger.info(f"Загружено {len(existing_products_by_ext_id)} существующих продуктов")
        
        # Обрабатываем батчами для оптимизации
        batch_size = 2000
        batch_num = 0
        for i in range(0, total_rows, batch_size):
            batch_num += 1
            batch = rows[i:i + batch_size]
            logger.info(f"Обрабатываем батч {batch_num} ({len(batch)} строк, строки {i+1}-{min(i+len(batch), total_rows)})")
            
            # Для первого батча показываем пример структуры данных
            if batch_num == 1 and batch:
                logger.info(f"Пример первой строки CSV: {list(batch[0].keys())}")
            
            # Коллекции для bulk операций
            brands_to_create = []
            products_to_create = []
            products_to_update = []
            snapshots_to_create = []
            
            # Временный кэш новых брендов в этом батче
            new_brands_cache = {}
            
            # Кэш для отслеживания дубликатов ext_id в текущем батче
            seen_ext_ids = set()
            
            with transaction.atomic():
                # Первый проход: собираем данные и создаем бренды
                parsed_rows = []
                for row in batch:
                    try:
                        # Извлекаем данные из CSV
                        code = row.get('Код', '').strip()
                        if not code:
                            skipped_no_code += 1
                            continue
                        
                        # Проверяем дубликат ext_id в текущем батче
                        if code in seen_ext_ids:
                            if len([err for err in errors if 'дубликат Код' in err]) < 5:
                                errors.append(f"Пропущен дубликат Код={code} в батче {batch_num}")
                            skipped_parse_errors += 1
                            continue
                        seen_ext_ids.add(code)
                        
                        part_number = row.get('Номенклатура', '').strip()
                        if not part_number:
                            skipped_no_part_number += 1
                            continue
                        
                        description = row.get('Описание', '').strip()
                        producer = row.get('Производитель', '').strip()
                        body_type = row.get('Тип корпуса', '').strip()
                        analogs = row.get('Аналоги', '').strip()
                        
                        # Цена 4 (может быть в формате с пробелами и запятыми)
                        price_ex_vat = None
                        price_str = row.get('Цена 4', '').strip()
                        if price_str:
                            try:
                                # Убираем все пробелы и неразрывные пробелы, заменяем запятую на точку
                                price_clean = price_str.replace(' ', '').replace('\xa0', '').replace(',', '.')
                                price_ex_vat = Decimal(price_clean)
                            except (InvalidOperation, ValueError) as e:
                                if len([err for err in errors if 'цены' in err]) < 5:
                                    errors.append(f"Ошибка парсинга цены '{price_str}' для Код={code}: {e}")
                                pass
                        
                        # Свободный остаток
                        stock_qty = 0
                        stock_str = row.get('Свободный остаток', '0').strip()
                        try:
                            # Убираем пробелы из числа перед конвертацией
                            stock_clean = stock_str.replace(' ', '').replace('\xa0', '') if stock_str else '0'
                            stock_qty = int(float(stock_clean)) if stock_clean else 0
                        except (ValueError, InvalidOperation):
                            pass
                        
                        # Определяем статус наличия
                        if stock_qty > 10:
                            stock_status = CompetitorPriceStockSnapshot.StockStatus.IN_STOCK
                        elif stock_qty > 0:
                            stock_status = CompetitorPriceStockSnapshot.StockStatus.LOW_STOCK
                        else:
                            stock_status = CompetitorPriceStockSnapshot.StockStatus.OUT_OF_STOCK
                        
                        # Создаем/находим бренд
                        brand = None
                        if producer:
                            # Проверяем в кэше существующих
                            if producer in existing_brands:
                                brand = existing_brands[producer]
                            # Проверяем в кэше новых в этом батче
                            elif producer in new_brands_cache:
                                brand = new_brands_cache[producer]
                            # Создаем новый бренд
                            else:
                                brand = CompetitorBrand(
                                    competitor=competitor,
                                    name=producer,
                                    ext_id=''
                                )
                                brands_to_create.append(brand)
                                new_brands_cache[producer] = brand
                        
                        # Сохраняем распарсенные данные
                        parsed_rows.append({
                            'code': code,
                            'part_number': part_number,
                            'description': description,
                            'brand': brand,
                            'producer': producer,
                            'body_type': body_type,
                            'analogs': analogs,
                            'price_ex_vat': price_ex_vat,
                            'stock_qty': stock_qty,
                            'stock_status': stock_status,
                            'row': row
                        })
                        
                    except Exception as e:
                        error_msg = f"Ошибка при парсинге строки {row.get('Код', 'unknown')}: {str(e)}"
                        logger.error(error_msg)
                        errors.append(error_msg)
                        skipped_parse_errors += 1
                        continue
                
                processed_rows += len(parsed_rows)
                
                logger.info(f"Распарсено {len(parsed_rows)} валидных строк из {len(batch)} в батче {batch_num}")
                
                # Создаем новые бренды батчем
                if brands_to_create:
                    created_brand_names = [b.name for b in brands_to_create]
                    CompetitorBrand.objects.bulk_create(brands_to_create, ignore_conflicts=True)
                    brands_created += len(brands_to_create)
                    logger.info(f"Создано {len(brands_to_create)} новых брендов")
                    
                    # Перезагружаем созданные бренды из БД для получения id
                    freshly_created_brands = CompetitorBrand.objects.filter(
                        competitor=competitor,
                        name__in=created_brand_names
                    ).select_related('competitor')
                    
                    # Обновляем кэш существующих брендов
                    for brand in freshly_created_brands:
                        existing_brands[brand.name] = brand
                        if brand.name in new_brands_cache:
                            new_brands_cache[brand.name] = brand
                
                # Второй проход: создаем/обновляем продукты
                for parsed in parsed_rows:
                    try:
                        existing_product = existing_products_by_ext_id.get(parsed['code'])
                        
                        # Получаем бренд из кэша (с актуальным id из БД)
                        brand = None
                        if parsed['producer']:
                            brand = existing_brands.get(parsed['producer'])
                        
                        if existing_product:
                            # Обновляем существующий продукт
                            existing_product.ext_id = parsed['code']
                            existing_product.part_number = parsed['part_number']
                            existing_product.name = parsed['description'] or parsed['part_number']
                            existing_product.brand = brand
                            existing_product.tech_params = {
                                'body_type': parsed['body_type'],
                                'analogs': parsed['analogs'],
                                'expected': parsed['row'].get('Ожидается', ''),
                                'shipment_multiplicity': parsed['row'].get('Кратность отгрузки', ''),
                            }
                            products_to_update.append(existing_product)
                            parsed['product'] = existing_product
                        else:
                            # Создаем новый продукт
                            new_product = CompetitorProduct(
                                competitor=competitor,
                                part_number=parsed['part_number'],
                                ext_id=parsed['code'],
                                name=parsed['description'] or parsed['part_number'],
                                brand=brand,
                                tech_params={
                                    'body_type': parsed['body_type'],
                                    'analogs': parsed['analogs'],
                                    'expected': parsed['row'].get('Ожидается', ''),
                                    'shipment_multiplicity': parsed['row'].get('Кратность отгрузки', ''),
                                }
                            )
                            products_to_create.append(new_product)
                            parsed['product'] = new_product
                            
                    except Exception as e:
                        error_msg = f"Ошибка при подготовке продукта {parsed['part_number']}: {str(e)}"
                        logger.error(error_msg)
                        errors.append(error_msg)
                        continue
                
                # Bulk create новых продуктов
                if products_to_create:
                    created_ext_ids = [p.ext_id for p in products_to_create]
                    CompetitorProduct.objects.bulk_create(products_to_create, ignore_conflicts=True)
                    products_created += len(products_to_create)
                    logger.info(f"Создано {len(products_to_create)} новых продуктов")
                    
                    # Перезагружаем созданные продукты из БД для получения id
                    freshly_created = CompetitorProduct.objects.filter(
                        competitor=competitor,
                        ext_id__in=created_ext_ids
                    ).select_related('brand', 'competitor')
                    
                    # Обновляем кэш существующих продуктов
                    for product in freshly_created:
                        existing_products_by_ext_id[product.ext_id] = product
                
                # Bulk update существующих продуктов
                if products_to_update:
                    CompetitorProduct.objects.bulk_update(
                        products_to_update,
                        ['ext_id', 'part_number', 'name', 'brand', 'tech_params', 'updated_at'],
                        batch_size=1000
                    )
                    products_updated += len(products_to_update)
                    logger.info(f"Обновлено {len(products_to_update)} продуктов")
                
                # Третий проход: создаем снимки
                for parsed in parsed_rows:
                    try:
                        if 'product' not in parsed:
                            continue
                        
                        # Используем продукт из кэша (с актуальным id из БД)
                        product = existing_products_by_ext_id.get(parsed['code'])
                        if not product:
                            logger.warning(f"Не найден продукт с ext_id={parsed['code']} для создания снимка")
                            continue
                        
                        snapshot = CompetitorPriceStockSnapshot(
                            competitor=competitor,
                            competitor_product=product,
                            collected_at=collected_at,
                            price_ex_vat=parsed['price_ex_vat'],
                            stock_qty=parsed['stock_qty'],
                            stock_status=parsed['stock_status'],
                            currency='USD',
                            raw_payload=dict(parsed['row'])
                        )
                        snapshots_to_create.append(snapshot)
                        
                    except Exception as e:
                        error_msg = f"Ошибка при создании снимка для {parsed.get('part_number', 'unknown')}: {str(e)}"
                        logger.error(error_msg)
                        errors.append(error_msg)
                        continue
                
                # Bulk create снимков
                if snapshots_to_create:
                    CompetitorPriceStockSnapshot.objects.bulk_create(
                        snapshots_to_create, 
                        ignore_conflicts=True,
                        batch_size=1000
                    )
                    snapshots_created += len(snapshots_to_create)
                    logger.info(f"Создано {len(snapshots_to_create)} снимков")
            
            logger.info(
                f"Батч завершен. Всего создано: продуктов={products_created}, "
                f"обновлено={products_updated}, брендов={brands_created}, снимков={snapshots_created}"
            )
        
        result = {
            "success": True,
            "total_rows": total_rows,
            "processed_rows": processed_rows,
            "products_created": products_created,
            "products_updated": products_updated,
            "brands_created": brands_created,
            "snapshots_created": snapshots_created,
            "skipped_no_code": skipped_no_code,
            "skipped_no_part_number": skipped_no_part_number,
            "skipped_parse_errors": skipped_parse_errors,
            "errors_count": len(errors),
            "errors": errors[:10] if errors else []  # Первые 10 ошибок
        }
        
        logger.info(
            f"✅ Импорт RCT завершен успешно! "
            f"Всего строк в CSV: {total_rows}, обработано валидных: {processed_rows} ({processed_rows/total_rows*100:.1f}% если total > 0). "
            f"Создано: {products_created} продуктов, {brands_created} брендов, {snapshots_created} снимков. "
            f"Обновлено: {products_updated} продуктов. "
            f"Пропущено: {skipped_no_code} без Кода, {skipped_no_part_number} без Номенклатуры, {skipped_parse_errors} с ошибками парсинга. "
            f"Всего ошибок: {len(errors)}"
        )
        return result
        
    except Exception as e:
        error_msg = f"Ошибка при парсинге CSV файла: {str(e)}"
        logger.error(error_msg, exc_info=True)
        return {"success": False, "error": error_msg}


@shared_task
def import_compel_from_https():
    """
    Импортирует данные о товарах конкурента COMPEL из HTTPS ZIP архива с DBF файлом.
    
    Оптимизированная версия с bulk-операциями:
    1. Находит конкурента COMPEL в базе
    2. Скачивает ZIP архив по HTTPS ссылке из data_url
    3. Распаковывает ZIP и извлекает DBF файл
    4. Парсит DBF со столбцами: CODE, PREFIX, NAME, PRODUCER, QTY, PRICE_1-PRICE_8, и т.д.
    5. Создает/обновляет записи CompetitorProduct батчами
    6. Создает снимки цен CompetitorPriceStockSnapshot батчами
    
    Маппинг полей:
    - CompetitorProduct.ext_id = CODE
    - CompetitorProduct.part_number = NAME
    - CompetitorProduct.brand = PRODUCER
    - CompetitorPriceStockSnapshot.price_ex_vat = PRICE_8 (если 0, то PRICE_7, ..., PRICE_1)
    - CompetitorPriceStockSnapshot.stock_qty = QTY
    - CompetitorPriceStockSnapshot.currency = USD
    
    Возвращает статистику по импортированным данным.
    """
    from django.conf import settings
    from django.utils import timezone
    
    logger.info("Начинаем импорт COMPEL из HTTPS")
    
    # Находим конкурента COMPEL в базе
    try:
        competitor = Competitor.objects.get(name="COMPEL", data_source_type=Competitor.DataSourceType.HTTPS)
    except Competitor.DoesNotExist:
        error_msg = "Конкурент COMPEL с типом HTTPS не найден в базе данных"
        logger.error(error_msg)
        return {"success": False, "error": error_msg}
    
    if not competitor.data_url:
        error_msg = "У конкурента COMPEL не указан URL для скачивания данных"
        logger.error(error_msg)
        return {"success": False, "error": error_msg}
    
    logger.info(f"Найден конкурент: {competitor.name}")
    logger.info(f"URL для скачивания: {competitor.data_url}")
    
    # Создаем директорию для скачивания
    base_download_dir = Path(settings.MEDIA_ROOT) / "https_downloads"
    competitor_dir = base_download_dir / competitor.name
    competitor_dir.mkdir(parents=True, exist_ok=True)
    
    zip_file_path = competitor_dir / "data.zip"
    extracted_dir = competitor_dir / "extracted"
    extracted_dir.mkdir(parents=True, exist_ok=True)
    
    try:
        # Скачиваем ZIP файл через HTTPS
        logger.info("Скачиваем ZIP архив...")
        
        # Настраиваем заголовки и аутентификацию если нужно
        headers = {}
        auth = None
        if competitor.username and competitor.password:
            auth = (competitor.username, competitor.password)
        
        response = requests.get(
            competitor.data_url,
            headers=headers,
            auth=auth,
            timeout=300,  # 5 минут таймаут
            stream=True
        )
        response.raise_for_status()
        
        # Сохраняем ZIP файл
        with open(zip_file_path, 'wb') as local_file:
            for chunk in response.iter_content(chunk_size=8192):
                local_file.write(chunk)
        
        logger.info(f"ZIP архив успешно скачан: {zip_file_path}")
        
        # Распаковываем ZIP архив
        logger.info("Распаковываем ZIP архив...")
        with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
            zip_ref.extractall(extracted_dir)
        
        logger.info(f"ZIP архив успешно распакован в: {extracted_dir}")
        
        # Ищем DBF файл в распакованной директории
        dbf_files = list(extracted_dir.glob("*.dbf")) + list(extracted_dir.glob("*.DBF"))
        if not dbf_files:
            error_msg = "DBF файл не найден в распакованном архиве"
            logger.error(error_msg)
            return {"success": False, "error": error_msg}
        
        dbf_file_path = dbf_files[0]
        logger.info(f"Найден DBF файл: {dbf_file_path}")
        
    except requests.exceptions.RequestException as e:
        error_msg = f"Ошибка при скачивании файла по HTTPS: {str(e)}"
        logger.error(error_msg)
        return {"success": False, "error": error_msg}
    except zipfile.BadZipFile as e:
        error_msg = f"Ошибка при распаковке ZIP архива: {str(e)}"
        logger.error(error_msg)
        return {"success": False, "error": error_msg}
    except Exception as e:
        error_msg = f"Ошибка при подготовке файлов: {str(e)}"
        logger.error(error_msg)
        return {"success": False, "error": error_msg}
    
    # Парсим DBF файл
    try:
        logger.info("Начинаем парсинг DBF файла...")
        
        products_created = 0
        products_updated = 0
        brands_created = 0
        snapshots_created = 0
        errors = []
        
        collected_at = timezone.now()
        
        # Читаем DBF файл с помощью simpledbf
        try:
            # Пробуем cp866 (стандартная кодировка для русских DBF)
            dbf = Dbf5(str(dbf_file_path), codec='cp866')
            df = dbf.to_dataframe()
            rows = df.to_dict('records')
        except Exception as e:
            # Пробуем другие кодировки
            logger.warning(f"Не удалось прочитать DBF с кодировкой cp866: {e}, пробуем cp1251")
            try:
                dbf = Dbf5(str(dbf_file_path), codec='cp1251')
                df = dbf.to_dataframe()
                rows = df.to_dict('records')
            except Exception as e2:
                logger.warning(f"Не удалось прочитать DBF с кодировкой cp1251: {e2}, пробуем utf-8")
                dbf = Dbf5(str(dbf_file_path), codec='utf-8')
                df = dbf.to_dataframe()
                rows = df.to_dict('records')
        
        total_rows = len(rows)
        logger.info(f"Прочитано {total_rows} строк из DBF файла")
        
        # Проверяем наличие необходимых колонок
        if rows:
            required_columns = ['CODE', 'NAME']
            available_columns = list(rows[0].keys())
            missing_columns = [col for col in required_columns if col not in available_columns]
            if missing_columns:
                error_msg = f"В DBF файле отсутствуют обязательные колонки: {missing_columns}. Доступные колонки: {available_columns}"
                logger.error(error_msg)
                return {"success": False, "error": error_msg}
            logger.info(f"Колонки DBF: {available_columns[:20]}{'...' if len(available_columns) > 20 else ''}")
        
        # Счетчики для пропущенных записей
        skipped_no_code = 0
        skipped_no_name = 0
        skipped_parse_errors = 0
        processed_rows = 0
        
        # Предзагрузка всех существующих брендов для кэширования
        logger.info("Загружаем существующие бренды...")
        existing_brands = {
            brand.name: brand 
            for brand in CompetitorBrand.objects.filter(competitor=competitor).select_related('competitor')
        }
        logger.info(f"Загружено {len(existing_brands)} существующих брендов")
        
        # Предзагрузка всех существующих продуктов для обновления
        logger.info("Загружаем существующие продукты...")
        existing_products_by_ext_id = {
            prod.ext_id: prod
            for prod in CompetitorProduct.objects.filter(competitor=competitor).select_related('brand', 'competitor')
        }
        logger.info(f"Загружено {len(existing_products_by_ext_id)} существующих продуктов")
        
        # Обрабатываем батчами для оптимизации
        batch_size = 2000
        batch_num = 0
        for i in range(0, total_rows, batch_size):
            batch_num += 1
            batch = rows[i:i + batch_size]
            logger.info(f"Обрабатываем батч {batch_num} ({len(batch)} строк, строки {i+1}-{min(i+len(batch), total_rows)})")
            
            # Для первого батча показываем пример структуры данных
            if batch_num == 1 and batch:
                logger.info(f"Пример первой строки DBF: {list(batch[0].keys())[:15]}...")
            
            # Коллекции для bulk операций
            brands_to_create = []
            products_to_create = []
            products_to_update = []
            snapshots_to_create = []
            
            # Временный кэш новых брендов в этом батче
            new_brands_cache = {}
            
            # Кэш для отслеживания дубликатов ext_id в текущем батче
            seen_ext_ids = set()
            
            with transaction.atomic():
                # Первый проход: собираем данные и создаем бренды
                parsed_rows = []
                for row in batch:
                    try:
                        # Извлекаем данные из DBF (обрабатываем pandas NaN)
                        code_val = row.get('CODE')
                        code = str(code_val).strip() if pd.notna(code_val) else ''
                        if not code:
                            skipped_no_code += 1
                            continue
                        
                        # Проверяем дубликат ext_id в текущем батче
                        if code in seen_ext_ids:
                            if len([err for err in errors if 'дубликат CODE' in err]) < 5:
                                errors.append(f"Пропущен дубликат CODE={code} в батче {batch_num}")
                            skipped_parse_errors += 1
                            continue
                        seen_ext_ids.add(code)
                        
                        name_val = row.get('NAME')
                        name = str(name_val).strip() if pd.notna(name_val) else ''
                        if not name:
                            skipped_no_name += 1
                            continue
                        
                        producer_val = row.get('PRODUCER')
                        producer = str(producer_val).strip() if pd.notna(producer_val) else ''
                        prefix_val = row.get('PREFIX')
                        prefix = str(prefix_val).strip() if pd.notna(prefix_val) else ''
                        
                        # Логика выбора цены: PRICE_8 -> PRICE_7 -> ... -> PRICE_1
                        price_ex_vat = None
                        for price_num in range(8, 0, -1):  # От 8 до 1
                            price_field = f'PRICE_{price_num}'
                            price_value = row.get(price_field)
                            
                            if pd.notna(price_value):
                                try:
                                    price_decimal = Decimal(str(price_value))
                                    if price_decimal > 0:
                                        price_ex_vat = price_decimal
                                        break  # Нашли ненулевую цену, прекращаем поиск
                                except (InvalidOperation, ValueError):
                                    pass
                        
                        # Количество на складе
                        stock_qty = 0
                        qty_value = row.get('QTY')
                        if pd.notna(qty_value):
                            try:
                                stock_qty = int(qty_value)
                            except (ValueError, TypeError):
                                pass
                        
                        # Определяем статус наличия
                        if stock_qty > 10:
                            stock_status = CompetitorPriceStockSnapshot.StockStatus.IN_STOCK
                        elif stock_qty > 0:
                            stock_status = CompetitorPriceStockSnapshot.StockStatus.LOW_STOCK
                        else:
                            stock_status = CompetitorPriceStockSnapshot.StockStatus.OUT_OF_STOCK
                        
                        # Создаем/находим бренд
                        brand = None
                        if producer:
                            # Проверяем в кэше существующих
                            if producer in existing_brands:
                                brand = existing_brands[producer]
                            # Проверяем в кэше новых в этом батче
                            elif producer in new_brands_cache:
                                brand = new_brands_cache[producer]
                            # Создаем новый бренд
                            else:
                                brand = CompetitorBrand(
                                    competitor=competitor,
                                    name=producer,
                                    ext_id=''
                                )
                                brands_to_create.append(brand)
                                new_brands_cache[producer] = brand
                        
                        # Сохраняем распарсенные данные
                        parsed_rows.append({
                            'code': code,
                            'name': name,
                            'prefix': prefix,
                            'brand': brand,
                            'producer': producer,
                            'price_ex_vat': price_ex_vat,
                            'stock_qty': stock_qty,
                            'stock_status': stock_status,
                            'row': row
                        })
                        
                    except Exception as e:
                        error_msg = f"Ошибка при парсинге строки {row.get('CODE', 'unknown')}: {str(e)}"
                        logger.error(error_msg)
                        errors.append(error_msg)
                        skipped_parse_errors += 1
                        continue
                
                processed_rows += len(parsed_rows)
                
                logger.info(f"Распарсено {len(parsed_rows)} валидных строк из {len(batch)} в батче {batch_num}")
                
                # Создаем новые бренды батчем
                if brands_to_create:
                    created_brand_names = [b.name for b in brands_to_create]
                    CompetitorBrand.objects.bulk_create(brands_to_create, ignore_conflicts=True)
                    brands_created += len(brands_to_create)
                    logger.info(f"Создано {len(brands_to_create)} новых брендов")
                    
                    # Перезагружаем созданные бренды из БД для получения id
                    freshly_created_brands = CompetitorBrand.objects.filter(
                        competitor=competitor,
                        name__in=created_brand_names
                    ).select_related('competitor')
                    
                    # Обновляем кэш существующих брендов
                    for brand in freshly_created_brands:
                        existing_brands[brand.name] = brand
                        if brand.name in new_brands_cache:
                            new_brands_cache[brand.name] = brand
                
                # Второй проход: создаем/обновляем продукты
                for parsed in parsed_rows:
                    try:
                        existing_product = existing_products_by_ext_id.get(parsed['code'])
                        
                        # Получаем бренд из кэша (с актуальным id из БД)
                        brand = None
                        if parsed['producer']:
                            brand = existing_brands.get(parsed['producer'])
                        
                        if existing_product:
                            # Обновляем существующий продукт
                            existing_product.ext_id = parsed['code']
                            existing_product.part_number = parsed['name']
                            existing_product.name = parsed['name']
                            existing_product.brand = brand
                            existing_product.tech_params = {
                                'prefix': parsed['prefix'],
                                'corpus': str(parsed['row'].get('CORPUS', '')) if pd.notna(parsed['row'].get('CORPUS')) else '',
                                'segment': str(parsed['row'].get('SEGMENT', '')) if pd.notna(parsed['row'].get('SEGMENT')) else '',
                                'sup_date': str(parsed['row'].get('SUP_DATE', '')) if pd.notna(parsed['row'].get('SUP_DATE')) else '',
                                'class_name': str(parsed['row'].get('CLASS_NAME', '')) if pd.notna(parsed['row'].get('CLASS_NAME')) else '',
                                'vendcode': str(parsed['row'].get('VENDCODE', '')) if pd.notna(parsed['row'].get('VENDCODE')) else '',
                                'weight': str(parsed['row'].get('WEIGHT', '')) if pd.notna(parsed['row'].get('WEIGHT')) else '',
                                'qnt_pack': str(parsed['row'].get('QNT_PACK', '')) if pd.notna(parsed['row'].get('QNT_PACK')) else '',
                                'moq': str(parsed['row'].get('MOQ', '')) if pd.notna(parsed['row'].get('MOQ')) else '',
                            }
                            products_to_update.append(existing_product)
                            parsed['product'] = existing_product
                        else:
                            # Создаем новый продукт
                            new_product = CompetitorProduct(
                                competitor=competitor,
                                part_number=parsed['name'],
                                ext_id=parsed['code'],
                                name=parsed['name'],
                                brand=brand,
                                tech_params={
                                    'prefix': parsed['prefix'],
                                    'corpus': str(parsed['row'].get('CORPUS', '')) if pd.notna(parsed['row'].get('CORPUS')) else '',
                                    'segment': str(parsed['row'].get('SEGMENT', '')) if pd.notna(parsed['row'].get('SEGMENT')) else '',
                                    'sup_date': str(parsed['row'].get('SUP_DATE', '')) if pd.notna(parsed['row'].get('SUP_DATE')) else '',
                                    'class_name': str(parsed['row'].get('CLASS_NAME', '')) if pd.notna(parsed['row'].get('CLASS_NAME')) else '',
                                    'vendcode': str(parsed['row'].get('VENDCODE', '')) if pd.notna(parsed['row'].get('VENDCODE')) else '',
                                    'weight': str(parsed['row'].get('WEIGHT', '')) if pd.notna(parsed['row'].get('WEIGHT')) else '',
                                    'qnt_pack': str(parsed['row'].get('QNT_PACK', '')) if pd.notna(parsed['row'].get('QNT_PACK')) else '',
                                    'moq': str(parsed['row'].get('MOQ', '')) if pd.notna(parsed['row'].get('MOQ')) else '',
                                }
                            )
                            products_to_create.append(new_product)
                            parsed['product'] = new_product
                            
                    except Exception as e:
                        error_msg = f"Ошибка при подготовке продукта {parsed['name']}: {str(e)}"
                        logger.error(error_msg)
                        errors.append(error_msg)
                        continue
                
                # Bulk create новых продуктов
                if products_to_create:
                    created_ext_ids = [p.ext_id for p in products_to_create]
                    CompetitorProduct.objects.bulk_create(products_to_create, ignore_conflicts=True)
                    products_created += len(products_to_create)
                    logger.info(f"Создано {len(products_to_create)} новых продуктов")
                    
                    # Перезагружаем созданные продукты из БД для получения id
                    freshly_created = CompetitorProduct.objects.filter(
                        competitor=competitor,
                        ext_id__in=created_ext_ids
                    ).select_related('brand', 'competitor')
                    
                    # Обновляем кэш существующих продуктов
                    for product in freshly_created:
                        existing_products_by_ext_id[product.ext_id] = product
                
                # Bulk update существующих продуктов
                if products_to_update:
                    CompetitorProduct.objects.bulk_update(
                        products_to_update,
                        ['ext_id', 'part_number', 'name', 'brand', 'tech_params', 'updated_at'],
                        batch_size=1000
                    )
                    products_updated += len(products_to_update)
                    logger.info(f"Обновлено {len(products_to_update)} продуктов")
                
                # Третий проход: создаем снимки
                for parsed in parsed_rows:
                    try:
                        if 'product' not in parsed:
                            continue
                        
                        # Используем продукт из кэша (с актуальным id из БД)
                        product = existing_products_by_ext_id.get(parsed['code'])
                        if not product:
                            logger.warning(f"Не найден продукт с ext_id={parsed['code']} для создания снимка")
                            continue
                        
                        snapshot = CompetitorPriceStockSnapshot(
                            competitor=competitor,
                            competitor_product=product,
                            collected_at=collected_at,
                            price_ex_vat=parsed['price_ex_vat'],
                            stock_qty=parsed['stock_qty'],
                            stock_status=parsed['stock_status'],
                            currency='USD',
                            raw_payload={
                                'CODE': str(parsed['row'].get('CODE', '')) if pd.notna(parsed['row'].get('CODE')) else '',
                                'PREFIX': str(parsed['row'].get('PREFIX', '')) if pd.notna(parsed['row'].get('PREFIX')) else '',
                                'NAME': str(parsed['row'].get('NAME', '')) if pd.notna(parsed['row'].get('NAME')) else '',
                                'PRODUCER': str(parsed['row'].get('PRODUCER', '')) if pd.notna(parsed['row'].get('PRODUCER')) else '',
                                'QTY': str(parsed['row'].get('QTY', '')) if pd.notna(parsed['row'].get('QTY')) else '',
                                'CORPUS': str(parsed['row'].get('CORPUS', '')) if pd.notna(parsed['row'].get('CORPUS')) else '',
                                'SEGMENT': str(parsed['row'].get('SEGMENT', '')) if pd.notna(parsed['row'].get('SEGMENT')) else '',
                            }
                        )
                        snapshots_to_create.append(snapshot)
                        
                    except Exception as e:
                        error_msg = f"Ошибка при создании снимка для {parsed.get('name', 'unknown')}: {str(e)}"
                        logger.error(error_msg)
                        errors.append(error_msg)
                        continue
                
                # Bulk create снимков
                if snapshots_to_create:
                    CompetitorPriceStockSnapshot.objects.bulk_create(
                        snapshots_to_create, 
                        ignore_conflicts=True,
                        batch_size=1000
                    )
                    snapshots_created += len(snapshots_to_create)
                    logger.info(f"Создано {len(snapshots_to_create)} снимков")
            
            logger.info(
                f"Батч завершен. Всего создано: продуктов={products_created}, "
                f"обновлено={products_updated}, брендов={brands_created}, снимков={snapshots_created}"
            )
        
        result = {
            "success": True,
            "total_rows": total_rows,
            "processed_rows": processed_rows,
            "products_created": products_created,
            "products_updated": products_updated,
            "brands_created": brands_created,
            "snapshots_created": snapshots_created,
            "skipped_no_code": skipped_no_code,
            "skipped_no_name": skipped_no_name,
            "skipped_parse_errors": skipped_parse_errors,
            "errors_count": len(errors),
            "errors": errors[:10] if errors else []  # Первые 10 ошибок
        }
        
        logger.info(
            f"✅ Импорт COMPEL завершен успешно! "
            f"Всего строк в DBF: {total_rows}, обработано валидных: {processed_rows} ({processed_rows/total_rows*100:.1f}% если total > 0). "
            f"Создано: {products_created} продуктов, {brands_created} брендов, {snapshots_created} снимков. "
            f"Обновлено: {products_updated} продуктов. "
            f"Пропущено: {skipped_no_code} без CODE, {skipped_no_name} без NAME, {skipped_parse_errors} с ошибками парсинга. "
            f"Всего ошибок: {len(errors)}"
        )
        return result
        
    except Exception as e:
        error_msg = f"Ошибка при парсинге DBF файла: {str(e)}"
        logger.error(error_msg, exc_info=True)
        return {"success": False, "error": error_msg}


@shared_task
def export_competitor_price_comparison_task():
    """
    Celery-задача для экспорта сравнения цен с конкурентами в Excel файл.
    ОПТИМИЗИРОВАННАЯ ВЕРСИЯ с предзагрузкой данных.
    
    Берёт все наши товары (part number) и находит точные совпадения
    по part_number у конкурентов, затем формирует таблицу сравнения цен.
    
    Returns:
        dict: Словарь с результатом экспорта и бинарным содержимым файла в base64
    """
    from collections import defaultdict
    from django.db.models import Prefetch, OuterRef, Subquery
    
    try:
        logger.info("Начинаем экспорт сравнения цен с конкурентами (оптимизированная версия)")
        
        # ШАГ 1: Предзагружаем все наши товары с последними ценами
        logger.info("Загружаем наши товары и последние цены...")
        our_products = list(Product.objects.select_related('brand').all())
        total_products = len(our_products)
        logger.info(f"Загружено {total_products} наших товаров")
        
        # ШАГ 2: Загружаем все последние цены одним запросом
        logger.info("Загружаем последние цены наших товаров...")
        product_ids = [p.id for p in our_products]
        
        # Получаем последние цены для каждого товара используя подзапрос
        latest_prices_subquery = OurPriceHistory.objects.filter(
            product_id=OuterRef('product_id')
        ).order_by('-moment').values('id')[:1]
        
        our_prices_dict = {}
        our_prices = OurPriceHistory.objects.filter(
            id__in=Subquery(latest_prices_subquery),
            product_id__in=product_ids
        ).select_related('product')
        
        for price in our_prices:
            our_prices_dict[price.product_id] = price
        
        logger.info(f"Загружено {len(our_prices_dict)} последних цен")
        
        # ШАГ 3: Загружаем последние данные нашего склада
        logger.info("Загружаем последние данные склада...")

        latest_stock_subquery = OurStockSnapshot.objects.filter(
            product_id=OuterRef('product_id')
        ).order_by('-moment').values('id')[:1]

        our_stock_dict = {}
        our_stock_snapshots = OurStockSnapshot.objects.filter(
            id__in=Subquery(latest_stock_subquery),
            product_id__in=product_ids,
        ).select_related('product')

        for stock_snapshot in our_stock_snapshots:
            our_stock_dict[stock_snapshot.product_id] = stock_snapshot

        logger.info(f"Загружено {len(our_stock_dict)} последних снимков склада")

        # ШАГ 4: Загружаем все товары конкурентов и группируем по part_number
        logger.info("Загружаем товары конкурентов...")
        competitor_products = CompetitorProduct.objects.select_related(
            'competitor', 'brand'
        ).all()
        
        # Группируем по part_number (нижний регистр для сравнения)
        comp_products_by_part = defaultdict(list)
        for comp_prod in competitor_products:
            comp_products_by_part[comp_prod.part_number.lower()].append(comp_prod)
        
        logger.info(f"Загружено {len(competitor_products)} товаров конкурентов, уникальных part_number: {len(comp_products_by_part)}")
        
        # ШАГ 5: Загружаем все последние snapshots для товаров конкурентов
        logger.info("Загружаем последние snapshots цен конкурентов...")
        comp_product_ids = [cp.id for cp in competitor_products]
        
        latest_snapshots_subquery = CompetitorPriceStockSnapshot.objects.filter(
            competitor_product_id=OuterRef('competitor_product_id')
        ).order_by('-collected_at').values('id')[:1]
        
        snapshots_dict = {}
        snapshots = CompetitorPriceStockSnapshot.objects.filter(
            id__in=Subquery(latest_snapshots_subquery),
            competitor_product_id__in=comp_product_ids
        ).select_related('competitor_product')
        
        for snapshot in snapshots:
            snapshots_dict[snapshot.competitor_product_id] = snapshot
        
        logger.info(f"Загружено {len(snapshots_dict)} последних snapshots")
        
        # ШАГ 6: Обрабатываем данные (теперь все данные в памяти!)
        logger.info("Начинаем обработку товаров...")
        comparison_data = []
        processed = 0
        matches_found = 0
        
        for product in our_products:
            processed += 1
            if processed % 1000 == 0:  # Увеличили до 1000 т.к. теперь быстро
                logger.info(f"Обработано {processed}/{total_products} товаров...")
            
            part_number = product.name
            
            # Получаем цену из предзагруженного словаря
            our_latest_price = our_prices_dict.get(product.id)
            
            our_price_inc_vat = None
            our_price_date = None
            our_stock_qty = None
            our_markup_percent = None
            our_cost_percent = None
            our_rmb_rate = None
            our_usd_rate = None
            
            if our_latest_price:
                # Рассчитываем цену с НДС
                if our_latest_price.price_ex_vat:
                    vat_multiplier = 1 + (float(our_latest_price.vat_rate) if our_latest_price.vat_rate else 0)
                    our_price_inc_vat = float(our_latest_price.price_ex_vat) * vat_multiplier
                our_price_date = our_latest_price.moment

            our_stock_snapshot = our_stock_dict.get(product.id)
            if our_stock_snapshot:
                if our_stock_snapshot.stock_qty is not None:
                    our_stock_qty = int(our_stock_snapshot.stock_qty)
                if our_stock_snapshot.markup_percent is not None:
                    our_markup_percent = round(float(our_stock_snapshot.markup_percent), 2)
                if our_stock_snapshot.cost_percent is not None:
                    our_cost_percent = round(float(our_stock_snapshot.cost_percent), 2)
                if our_stock_snapshot.rmb_rate is not None:
                    our_rmb_rate = round(float(our_stock_snapshot.rmb_rate), 4)
                if our_stock_snapshot.usd_rate is not None:
                    our_usd_rate = round(float(our_stock_snapshot.usd_rate), 4)
            
            # Ищем совпадения в предзагруженном словаре
            comp_products = comp_products_by_part.get(part_number.lower(), [])
            
            if comp_products:
                matches_found += 1
                # Обрабатываем каждый товар конкурента
                for comp_product in comp_products:
                    # Получаем snapshot из предзагруженного словаря
                    latest_snapshot = snapshots_dict.get(comp_product.id)
                    
                    comp_price_ex_vat = None
                    comp_stock_qty = None
                    comp_price_date = None
                    price_difference = None
                    price_difference_pct = None
                    
                    if latest_snapshot:
                        comp_price_ex_vat = float(latest_snapshot.price_ex_vat) if latest_snapshot.price_ex_vat else None
                        comp_stock_qty = latest_snapshot.stock_qty
                        comp_price_date = latest_snapshot.collected_at
                        
                        # Вычисляем разницу в ценах (наша цена с НДС - цена конкурента без НДС)
                        if our_price_inc_vat and comp_price_ex_vat:
                            price_difference = our_price_inc_vat - comp_price_ex_vat
                            price_difference_pct = (price_difference / comp_price_ex_vat) * 100
                    
                    comparison_data.append({
                        'Part Number': part_number,
                        'Наш бренд': product.brand.name if product.brand else '',
                        'Наша цена': our_price_inc_vat,
                        'Курс юаня': our_rmb_rate,
                        'Курс доллара': our_usd_rate,
                        'Дата нашей цены': our_price_date.strftime('%Y-%m-%d %H:%M') if our_price_date else '',
                        'Наш остаток': our_stock_qty,
                        'Наша наценка (%)': our_markup_percent,
                        'Наши затраты (%)': our_cost_percent,
                        'Конкурент': comp_product.competitor.name,
                        'Бренд конкурента': comp_product.brand.name if comp_product.brand else '',
                        'Цена конкурента': comp_price_ex_vat,
                        'Дата цены конкурента': comp_price_date.strftime('%Y-%m-%d %H:%M') if comp_price_date else '',
                        'Остаток у конкурента': comp_stock_qty,
                        'Разница в цене': round(price_difference, 2) if price_difference is not None else None,
                        'Разница в цене (%)': round(price_difference_pct, 2) if price_difference_pct is not None else None,
                    })
            else:
                # Товар без совпадений у конкурентов
                comparison_data.append({
                    'Part Number': part_number,
                    'Наш бренд': product.brand.name if product.brand else '',
                    'Наша цена': our_price_inc_vat,
                    'Курс юаня': our_rmb_rate,
                    'Курс доллара': our_usd_rate,
                    'Дата нашей цены': our_price_date.strftime('%Y-%m-%d %H:%M') if our_price_date else '',
                    'Наш остаток': our_stock_qty,
                    'Наша наценка (%)': our_markup_percent,
                    'Наши затраты (%)': our_cost_percent,
                    'Конкурент': 'Нет совпадений',
                    'Бренд конкурента': '',
                    'Цена конкурента': None,
                    'Дата цены конкурента': '',
                    'Остаток у конкурента': None,
                    'Разница в цене': None,
                    'Разница в цене (%)': None,
                })
        
        logger.info(f"Обработка завершена. Всего товаров: {total_products}, найдено совпадений: {matches_found}")
        
        # Создаём DataFrame
        df = pd.DataFrame(comparison_data)
        
        # Создаём Excel файл в памяти
        excel_buffer = BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Сравнение цен', index=False)
            
            # Получаем workbook и worksheet для форматирования
            workbook = writer.book
            worksheet = writer.sheets['Сравнение цен']
            
            # Автоподбор ширины столбцов
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Максимум 50 символов
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Закрепляем первую строку (заголовки)
            worksheet.freeze_panes = 'A2'
        
        # Получаем содержимое буфера
        excel_buffer.seek(0)
        file_data = excel_buffer.getvalue()
        
        # Преобразуем бинарные данные в base64 для безопасной передачи через JSON
        encoded_data = base64.b64encode(file_data).decode('utf-8')
        
        # Формируем результат
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"price_comparison_{timestamp}.xlsx"
        
        logger.info(f"✅ Excel файл создан успешно: {filename}, записей: {len(comparison_data)}")
        
        return {
            'success': True,
            'filename': filename,
            'data': encoded_data,
            'records': len(comparison_data),
            'total_products': total_products,
            'matches_found': matches_found
        }
        
    except Exception as e:
        logger.error(f"❌ Ошибка при экспорте сравнения цен: {str(e)}", exc_info=True)
        return {
            'success': False,
            'error': f"Ошибка при экспорте: {str(e)}"
        }


@shared_task
def export_competitor_sales_task(date_from=None, date_to=None, competitor_ids=None):
    """
    Celery-задача для экспорта продаж конкурентов в Excel файл.
    ОПТИМИЗИРОВАННАЯ ВЕРСИЯ с использованием агрегации Django ORM.
    
    Анализирует изменения в остатках товаров конкурентов за период 
    и рассчитывает предполагаемые продажи (уменьшение остатка).
    
    Parameters:
        date_from (str): Начальная дата в формате YYYY-MM-DD (по умолчанию: 30 дней назад)
        date_to (str): Конечная дата в формате YYYY-MM-DD (по умолчанию: сегодня)
        competitor_ids (list): Список ID конкурентов для анализа (по умолчанию: все конкуренты)
    
    Returns:
        dict: Словарь с результатом экспорта и бинарным содержимым файла в base64
    """
    from datetime import datetime, timedelta
    from django.utils import timezone
    from django.db.models import Min, Max, Avg, Count, F, Q, Subquery, OuterRef
    
    try:
        logger.info("Начинаем экспорт продаж конкурентов (оптимизированная версия)")
        
        # Устанавливаем период по умолчанию если не указан
        if not date_to:
            end_date = timezone.now()
        else:
            end_date = timezone.make_aware(datetime.strptime(date_to, '%Y-%m-%d'))
        
        if not date_from:
            start_date = end_date - timedelta(days=30)
        else:
            start_date = timezone.make_aware(datetime.strptime(date_from, '%Y-%m-%d'))
        
        logger.info(f"Период анализа: с {start_date.date()} по {end_date.date()}")
        
        # Логируем выбранных конкурентов
        if competitor_ids:
            logger.info(f"Фильтр по конкурентам: {competitor_ids}")
            # Получаем названия конкурентов для отчета
            selected_competitors = Competitor.objects.filter(id__in=competitor_ids).values_list('name', flat=True)
            competitors_info = ', '.join(selected_competitors)
            logger.info(f"Выбранные конкуренты: {competitors_info}")
        else:
            logger.info("Анализ всех конкурентов")
            competitors_info = "Все конкуренты"
        
        # ШАГ 1: Получаем уникальные товары конкурентов, у которых есть снимки за период
        logger.info("Получаем список товаров с данными за период...")
        snapshots_query = CompetitorPriceStockSnapshot.objects.filter(
            collected_at__gte=start_date,
            collected_at__lte=end_date
        )
        
        # Применяем фильтр по конкурентам если указан
        if competitor_ids:
            snapshots_query = snapshots_query.filter(competitor_id__in=competitor_ids)
        
        products_with_snapshots = snapshots_query.values('competitor_product_id').distinct()
        
        product_ids = [p['competitor_product_id'] for p in products_with_snapshots]
        total_products = len(product_ids)
        
        logger.info(f"Найдено {total_products} уникальных товаров конкурентов")
        
        if total_products == 0:
            logger.warning("Нет данных за указанный период")
            return {
                'success': False,
                'error': 'Нет данных за указанный период'
            }
        
        # ШАГ 2: Обрабатываем товары пакетами для экономии памяти
        logger.info("Начинаем обработку товаров пакетами...")
        sales_data = []
        BATCH_SIZE = 100  # Обрабатываем по 100 товаров за раз
        processed = 0
        
        for i in range(0, total_products, BATCH_SIZE):
            batch_ids = product_ids[i:i + BATCH_SIZE]
            logger.info(f"Обработка пакета {i // BATCH_SIZE + 1}/{(total_products + BATCH_SIZE - 1) // BATCH_SIZE}...")
            
            # Базовый queryset для подзапросов
            base_snapshot_filter = {
                'competitor_product_id': OuterRef('competitor_product_id'),
                'collected_at__gte': start_date,
                'collected_at__lte': end_date
            }
            
            # Получаем первые снимки для каждого товара в пакете
            first_snapshots_subquery = CompetitorPriceStockSnapshot.objects.filter(
                **base_snapshot_filter
            ).order_by('collected_at').values('id')[:1]
            
            # Получаем последние снимки для каждого товара в пакете
            last_snapshots_subquery = CompetitorPriceStockSnapshot.objects.filter(
                **base_snapshot_filter
            ).order_by('-collected_at').values('id')[:1]
            
            # Загружаем только нужные снимки для этого пакета
            batch_snapshots_query = CompetitorPriceStockSnapshot.objects.filter(
                Q(id__in=Subquery(first_snapshots_subquery)) | Q(id__in=Subquery(last_snapshots_subquery)),
                competitor_product_id__in=batch_ids
            )
            
            # Применяем фильтр по конкурентам если указан
            if competitor_ids:
                batch_snapshots_query = batch_snapshots_query.filter(competitor_id__in=competitor_ids)
            
            batch_snapshots = batch_snapshots_query.select_related(
                'competitor',
                'competitor_product',
                'competitor_product__brand'
            ).order_by('competitor_product_id', 'collected_at')
            
            # Группируем снимки по товарам для этого пакета
            from collections import defaultdict
            product_snapshots_batch = defaultdict(list)
            for snapshot in batch_snapshots:
                product_snapshots_batch[snapshot.competitor_product_id].append(snapshot)
            
            # Анализируем каждый товар в пакете
            for product_id, snapshots_list in product_snapshots_batch.items():
                processed += 1
                
                if len(snapshots_list) < 2:
                    # Нужно минимум 2 снимка для сравнения
                    continue
                
                # Сортируем по дате
                snapshots_list.sort(key=lambda x: x.collected_at)
                
                # Берем первый и последний снимки
                first_snapshot = snapshots_list[0]
                last_snapshot = snapshots_list[-1]
                
                # Получаем данные о товаре
                competitor_product = first_snapshot.competitor_product
                competitor_name = first_snapshot.competitor.name
                part_number = competitor_product.part_number
                brand_name = competitor_product.brand.name if competitor_product.brand else ''
                
                # Рассчитываем изменение остатка
                first_stock = first_snapshot.stock_qty if first_snapshot.stock_qty is not None else 0
                last_stock = last_snapshot.stock_qty if last_snapshot.stock_qty is not None else 0
                stock_change = first_stock - last_stock
                
                # Если остаток уменьшился - считаем это продажами
                sold_qty = max(0, stock_change)
                
                # Получаем среднюю цену
                prices = [s.price_ex_vat for s in snapshots_list if s.price_ex_vat is not None]
                avg_price = sum(prices) / len(prices) if prices else None
                
                # Цена последнего снимка
                last_price = last_snapshot.price_ex_vat
                
                # Рассчитываем сумму продаж
                sales_amount = None
                if sold_qty > 0 and avg_price:
                    sales_amount = sold_qty * float(avg_price)
                
                # Добавляем запись
                sales_data.append({
                    'Конкурент': competitor_name,
                    'Part Number': part_number,
                    'Бренд': brand_name,
                    'Остаток на начало': first_stock,
                    'Остаток на конец': last_stock,
                    'Продано (шт)': sold_qty,
                    'Средняя цена': round(float(avg_price), 2) if avg_price else None,
                    'Цена актуальная': round(float(last_price), 2) if last_price else None,
                    'Сумма продаж': round(sales_amount, 2) if sales_amount else None,
                    'Дата первого снимка': first_snapshot.collected_at.strftime('%Y-%m-%d'),
                    'Дата последнего снимка': last_snapshot.collected_at.strftime('%Y-%m-%d'),
                    'Количество снимков': len(snapshots_list),
                })
            
            # Очищаем память после обработки пакета
            del batch_snapshots
            del product_snapshots_batch
        
        logger.info(f"Анализ завершен. Обработано товаров: {total_products}")
        
        # Сортируем по сумме продаж (по убыванию)
        sales_data.sort(key=lambda x: x['Сумма продаж'] if x['Сумма продаж'] else 0, reverse=True)
        
        # Создаём DataFrame
        df = pd.DataFrame(sales_data)
        
        # Добавляем итоговые строки
        total_sold = df['Продано (шт)'].sum()
        total_sales_amount = df['Сумма продаж'].sum()
        
        # Создаём Excel файл в памяти
        excel_buffer = BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Продажи конкурентов', index=False, startrow=1)
            
            # Получаем workbook и worksheet для форматирования
            workbook = writer.book
            worksheet = writer.sheets['Продажи конкурентов']
            
            # Добавляем заголовок с периодом и фильтрами
            header_text = f"Анализ продаж конкурентов за период: {start_date.strftime('%Y-%m-%d')} - {end_date.strftime('%Y-%m-%d')}"
            if competitor_ids:
                header_text += f" | Конкуренты: {competitors_info}"
            worksheet.cell(row=1, column=1).value = header_text
            
            # Автоподбор ширины столбцов
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Закрепляем первые две строки (заголовок + названия столбцов)
            worksheet.freeze_panes = 'A3'
            
            # Добавляем итоговую строку
            last_row = len(df) + 3
            worksheet.cell(row=last_row, column=1).value = "ИТОГО:"
            worksheet.cell(row=last_row, column=6).value = total_sold
            worksheet.cell(row=last_row, column=9).value = round(total_sales_amount, 2) if total_sales_amount else 0
            
            # Делаем итоговую строку жирной
            from openpyxl.styles import Font
            for col in range(1, 13):
                cell = worksheet.cell(row=last_row, column=col)
                cell.font = Font(bold=True)
        
        # Получаем содержимое буфера
        excel_buffer.seek(0)
        file_data = excel_buffer.getvalue()
        
        # Преобразуем бинарные данные в base64
        encoded_data = base64.b64encode(file_data).decode('utf-8')
        
        # Формируем результат
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"competitor_sales_{timestamp}.xlsx"
        
        logger.info(f"✅ Excel файл создан успешно: {filename}, записей: {len(sales_data)}")
        logger.info(f"   Всего продано единиц: {total_sold}, на сумму: {round(total_sales_amount, 2) if total_sales_amount else 0}")
        
        return {
            'success': True,
            'filename': filename,
            'data': encoded_data,
            'records': len(sales_data),
            'total_sold': int(total_sold),
            'total_sales_amount': round(float(total_sales_amount), 2) if total_sales_amount else 0,
            'period_from': start_date.strftime('%Y-%m-%d'),
            'period_to': end_date.strftime('%Y-%m-%d')
        }
        
    except Exception as e:
        logger.error(f"❌ Ошибка при экспорте продаж конкурентов: {str(e)}", exc_info=True)
        return {
            'success': False,
            'error': f"Ошибка при экспорте: {str(e)}"
        }
